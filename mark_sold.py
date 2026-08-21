import json
import os
from datetime import datetime, timezone
import sys
import re
import argparse

_REPO_DIR = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(_REPO_DIR, 'coins.json')

def parse_sales_text(text):
    """
    Robustly parses sales strings in multiple formats:
      - 'id 105 2000 quique' or '1088 $4.000 Seba Verna' (<ID> <PRICE> <CLIENTS>)
      - 'id 1089 bruni 18000' or '1089 bruni $18.000' (<ID> <CLIENTS> <PRICE>)
      - Comma-separated: '1090 $3.000 Alegre German, Seba Verna'
      - Multi-line sales blocks
    Returns a dict mapping coin_id -> requested_quantity.
    """
    cleaned = re.sub(r'^(?:pe|mv|publicar[\s\-_]excel|asentar|registrar|venta[s]?)\s*[:,\-]?\s*', '', text.strip(), flags=re.IGNORECASE)
    counts = {}
    lines = [l.strip() for l in cleaned.split('\n') if l.strip()]

    # Try loading known clients from VTAS.xlsx for accurate counting
    known_clients = []
    try:
        import openpyxl
        vtas_path = os.path.join(os.path.expanduser('~'), 'Desktop', 'Popper', 'VTAS.xlsx')
        if os.path.exists(vtas_path):
            wb = openpyxl.load_workbook(vtas_path, data_only=True)
            ws = wb['Clientes']
            known_clients = [str(ws.cell(row=r, column=1).value).strip().lower() for r in range(2, ws.max_row+1) if ws.cell(row=r, column=1).value]
    except Exception:
        pass

    for line in lines:
        line_clean = re.sub(r'\([^\)]*el resto no[^\)]*\)', '', line, flags=re.IGNORECASE).strip()
        line_clean = re.sub(r'\b(?:pe\s+y\s+mv|pe\s+y\s+mark[\s\-_]sold|mv\s+y\s+pe|pe|mv)\s*$', '', line_clean, flags=re.IGNORECASE).strip()
        if not line_clean:
            continue

        m1 = re.match(r'^(?:id\s+)?(?P<id>\d+)\s+(?P<price>[$]?\d+(?:[\.,]\d+)?(?:\s*(?:usd|dolar(?:es)?))?)\s+(?P<clients>.+)$', line_clean, re.IGNORECASE)
        m2 = re.match(r'^(?:id\s+)?(?P<id>\d+)\s+(?P<clients>.+?)\s+(?P<price>[$]?\d+(?:[\.,]\d+)?(?:\s*(?:usd|dolar(?:es)?))?)$', line_clean, re.IGNORECASE)

        m = m1 or m2
        if m:
            coin_id = int(m.group('id'))
            clients_raw = m.group('clients').strip()
            if ',' in clients_raw:
                qty = len([c for c in clients_raw.split(',') if c.strip()])
            else:
                tokens = clients_raw.split()
                if known_clients:
                    idx = 0
                    buyer_count = 0
                    while idx < len(tokens):
                        found = False
                        for span in [3, 2, 1]:
                            if idx + span <= len(tokens):
                                chunk = ' '.join(tokens[idx:idx+span]).lower()
                                if any(chunk in k for k in known_clients):
                                    buyer_count += 1
                                    idx += span
                                    found = True
                                    break
                        if not found:
                            buyer_count += 1
                            idx += 1
                    qty = max(1, buyer_count)
                else:
                    qty = max(1, len(tokens))
            counts[coin_id] = counts.get(coin_id, 0) + qty
        else:
            # Sub pattern matching
            pattern = re.compile(
                r'(?:id\s+)?(?P<id>\d+)\s+'
                r'(?P<price>(?:\$)?\d+(?:[\.,]\d+)?(?:\s*(?:usd|dolar(?:es)?))?)\s+'
                r'(?P<clients>.+?)'
                r'(?=(?:\s+(?:id\s+)?\d+\s+(?:\$)?\d+)|$)',
                re.IGNORECASE
            )
            matches = list(pattern.finditer(line_clean))
            if matches:
                for match_item in matches:
                    cid = int(match_item.group('id'))
                    c_raw = match_item.group('clients').strip()
                    toks = c_raw.split()
                    counts[cid] = counts.get(cid, 0) + max(1, len(toks))
            else:
                id_matches = re.findall(r'\b(?:id\s*)?(\d+)\b', line_clean, re.IGNORECASE)
                for mid in id_matches:
                    cid = int(mid)
                    counts[cid] = counts.get(cid, 0) + 1

    return counts

def parse_arguments():
    parser = argparse.ArgumentParser(description="Mark coins as sold / decrement stock in coins.json")
    parser.add_argument("ids", nargs="*", help="List of IDs or ID:QTY (e.g. 105 2034 2034 or 2034:2)")
    parser.add_argument("--text", default=None, help="Raw sales text e.g. 'id 105 2000 quique' or multiline list")
    parser.add_argument("--json", dest="json_input", default=None, help="JSON list of {id, qty}")
    parser.add_argument("--skip-stock-check", action="store_true", help="Allow selling down to 0 even if requested > initial stock")
    return parser.parse_args()

def main():
    args = parse_arguments()
    requested_counts = {}
    
    if args.text:
        requested_counts = parse_sales_text(args.text)
            
    elif args.json_input:
        data = json.loads(args.json_input)
        for item in data:
            cid = int(item['id'])
            qty = int(item.get('qty', item.get('cantidad', 1)))
            requested_counts[cid] = requested_counts.get(cid, 0) + qty
            
    elif args.ids:
        for arg in args.ids:
            if ":" in arg:
                parts = arg.split(":")
                cid = int(parts[0])
                qty = int(parts[1])
                requested_counts[cid] = requested_counts.get(cid, 0) + qty
            else:
                try:
                    cid = int(arg)
                    requested_counts[cid] = requested_counts.get(cid, 0) + 1
                except ValueError:
                    # Might be passed like 'id 105'
                    pass
    else:
        print("Usage: python3 mark_sold.py <id1> <id2> ... or --text 'id 105 2000 quique'")
        sys.exit(1)
        
    if not requested_counts:
        print("ERROR: No valid coin IDs provided")
        sys.exit(1)

    if not os.path.exists(file_path):
        print(f"ERROR: {file_path} not found")
        sys.exit(1)

    with open(file_path, 'r', encoding='utf-8') as f:
        coins = json.load(f)

    # ─────────────────────────────────────────────────────────────
    # DOBLE VERIFICACIÓN DE EXISTENCIA Y STOCK
    # ─────────────────────────────────────────────────────────────
    coin_map = {c['id']: c for c in coins}
    verification_errors = []

    for cid, req_qty in requested_counts.items():
        if cid not in coin_map:
            verification_errors.append(f"Moneda ID {cid} NO existe en coins.json.")
            continue
            
        coin = coin_map[cid]
        title = coin.get('title', 'Sin título')
        status = coin.get('status', '')
        
        if status == 'sold':
            current_stock = 0
        else:
            current_stock = int(coin.get('cantidad', 1))
            
        if not args.skip_stock_check and current_stock - req_qty < 0:
            verification_errors.append(
                f"Stock insuficiente para ID {cid} ({title}). Stock disponible: {current_stock}, Solicitado: {req_qty}. El stock nunca puede ser menor a 0."
            )

    if verification_errors:
        print("ERROR_DOBLE_VERIFICACION_STOCK:")
        for err in verification_errors:
            print(f"  - {err}")
        sys.exit(1)

    # ─────────────────────────────────────────────────────────────
    # APLICACIÓN DE CAMBIOS (STOCK / SOLD)
    # ─────────────────────────────────────────────────────────────
    updated_coins = []
    now_iso = datetime.now(timezone.utc).isoformat()

    for cid, req_qty in requested_counts.items():
        coin = coin_map[cid]
        if coin.get('status') == 'sold':
            current_stock = 0
        else:
            current_stock = int(coin.get('cantidad', 1))
            
        new_stock = max(0, current_stock - req_qty)
        
        if new_stock == 0:
            coin['status'] = 'sold'
            coin['soldAt'] = now_iso
            if 'cantidad' in coin:
                coin['cantidad'] = 0
            updated_coins.append((coin, True, 0, req_qty))
        else:
            coin['cantidad'] = new_stock
            if 'status' in coin and coin['status'] == 'sold':
                del coin['status']
            updated_coins.append((coin, False, new_stock, req_qty))

    # Guardar coins.json
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(coins, f, indent=2, ensure_ascii=False)

    # No se escribe en operational_log.md: esa bitácora guarda lógica, no ventas.
    # La venta ya queda registrada en coins.json, en VTAS.xlsx y en el historial de git.

    ids_str = ",".join(str(c['id']) for c, _, _, _ in updated_coins)
    titles_str = "; ".join(c['title'] for c, _, _, _ in updated_coins)
    print(f"SUCCESS:{ids_str}:{titles_str}")
    
    for found_coin, became_sold, rem_stock, req_qty in updated_coins:
        cid = found_coin.get('id')
        title = found_coin.get('title')
        if became_sold:
            print(f"ID {cid}: {title} → VENDIDO (status: sold, stock: 0)")
        else:
            print(f"ID {cid}: {title} → STOCK REDUCIDO (vendidas: {req_qty}, restante: {rem_stock})")

if __name__ == '__main__':
    main()
